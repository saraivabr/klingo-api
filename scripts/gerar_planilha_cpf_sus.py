"""
Gera planilha Excel com CPF e Cartão SUS dos atendimentos de fevereiro/2026.
"""
import sys
sys.path.insert(0, '/Users/saraiva/Documents/IRB')
from klingo_api import KlingoAPI
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime

api = KlingoAPI(domain="irb")
api.login("FELLIPE.SARAIVA", "FELLIPE.SARAIVA1")

atendimentos = [
    # Já têm dados no relatório original (CPF e SUS preenchidos)
    (3697, "MARIA APARECIDA FERREIRA DE SOUZA", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3700, "VICTOR RAGONETE NARDUCI", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3704, "ANDRE DE CAMPOS CORDEIRO", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3709, "CARLOS FERREIRA NETO", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3715, "MIGUEL BALIONE DA SILVA", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3720, "RITA GOMES DE OLIVEIRA", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3723, "LEONIRA RIBAS DOS REIS", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3732, "MARISOL MANRIQUE COTRINA", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3812, "JULIANA TENORIO", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3813, "JULIA VITORIA SILVA CAVALCANTE", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3814, "VITOR GOMES DA SILVA", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3815, "GIAN CARLO DUARTE BORTOLUZZI", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3816, "EMANUEL NOLASCO SILVA", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3817, "PEDRO PAULO ALEXOPULOS DOS SANTOS", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3818, "MARCELY NASCIMENTO TAVARES", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3819, "GABRIELI EDUARDA OLIVEIRA FARIAS", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3820, "EMANUEL PAZ FRANCO", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3821, "GUSTAVO CORREA MELO", "12/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    # Com CPF/SUS faltando (buscados da API)
    (3699, "ANTONIO DE SOUZA", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3718, "APARECIDO JOSE DA SILVA", "11/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "Particular"),
    (3851, "REBECA MOREIRA DA SILVA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3856, "HADASSA MOREIRA DA SILVA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3859, "DAVI RAACH BRANDAO PASSOS", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3860, "OLIVIA GABRIELLY DA SILVA MANFROI", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3861, "LAURA BUENO DE OLIVEIRA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3865, "LAIS BUENO DE OLIVEIRA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3866, "ETHAN GABRIEL OLIVEIRA RAMOS RODRIGUES", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3867, "MURILLO FERREIRA ROCHA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3869, "HELENNA FERREIRA PRADO", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3871, "MARIA HELOISA CARDOSO SCATOLIM", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3872, "MAITE AGUIAR PINHEIRO", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3873, "NOAH MARTINS GOMES", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3876, "ARTHUR GABRYEL ASSUNCAO SOUZA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3878, "ARTHUR VINICIUS DE SOUZA FURTADO", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3881, "JULIA BEATRIZ FERREIRA SILVA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3884, "JOVANE KAUE PIRES DA SILVA", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3887, "KEVIN RYAN MARTINS DOS SANTOS", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3888, "ALICE FERREIRA NASCIMENTO", "13/02/2026", "Pediatria", "DRA BEATRIZ ANTUNES", "Particular"),
    (3932, "VALDIR VENSON", "13/02/2026", "Gastroenterologia", "DRA ANA TRAJANO", "Particular"),
    (3933, "KAYKI SACCHETTI DE OLIVEIRA", "13/02/2026", "Gastroenterologia", "DRA ANA TRAJANO", "Particular"),
    (3934, "NEUZA MIRANDA GONCALVES", "13/02/2026", "Gastroenterologia", "DRA ANA TRAJANO", "Particular"),
    (3935, "ANGELINA SAURIN", "13/02/2026", "Gastroenterologia", "DRA ANA TRAJANO", "Particular"),
    (3946, "MARCOS EMANUEL DOS SANTOS ALVARINTHO", "19/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3947, "GENILTON MACIEL MUNIZ", "19/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3948, "IRENE MENDES DOS SANTOS", "19/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3949, "CARLOS ROBERTO MACAUBAS", "19/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3950, "JEREMIAS ALVES DA SILVA", "19/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3951, "IVANETE RODRIGUES", "19/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (3957, "DIEGO RODRIGUES", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3959, "ROSANGELA ALVES BORBA", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3960, "BRUNO JANEIRO DA SILVA", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3962, "FABIO JUNIOR DE CARVALHO", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3964, "DAVI RAFAEL MACHADO SILVA", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3965, "DEBORAH KETELLYN CARDOSO LOPES", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3966, "IGOR HENRIQUE OLIVEIRA SANTOS", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3967, "JOHN MICHAEL MATHEUS ARIAS", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3968, "YORHANE MACEDO ARRIGO", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3969, "ELIANE APARECIDA GOMES GONCALVES", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3971, "JOZIANE MARA MACEDO DA SILVA", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3972, "JOSILDO CLAUDINO DE MORAIS", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (3973, "JOSUE ALEXANDER ROSILLO CARTAYA", "19/02/2026", "Psiquiatria", "DRA PATRICIA BENTO", "Particular"),
    (4115, "NICOMERIA MEDEIROS SCHWENCK NETA", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "Particular"),
    (4117, "MAYCON CEZAR BRAGANCA TAVARES", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "Particular"),
    (4118, "MATHEUS DE LIMA SOUZA", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "Particular"),
    (4119, "RODRIGO HENRIQUE VERLY DA SILVA", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "Particular"),
    (4121, "MARIA HELENA DA SILVA", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "Particular"),
    (4126, "ELISIANE FERREIRA", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "Particular"),
    (4134, "DEOCLECIO CORREIA PEDROSO", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "Particular"),
    (4120, "TALIA SOUZA DE OLIVEIRA", "23/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (4127, "GABRIEL MATHEUS DOS SANTOS", "23/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (4129, "ANY IZABELY LEONEL DE SOUZA", "23/02/2026", "Neurologia", "DR ANGELO CAMPOS", "Particular"),
    (4135, "ADILSON ALVES ALMEIDA", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4136, "ZELIA MARIA DEMEDEIROS", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4139, "ADILCE GOMES DOSSANTOS", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4140, "NATALIA CLECIANE DE PAULA SOLIS", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "SUS"),
    (4141, "ROSA ROCHA DESOUZA", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4142, "TANIA JUDITE MIOTTI", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "SUS"),
    (4144, "SIRLEIDE PARENTE DESOUZA", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4145, "JOSILDO CLAUDINO DE MORAIS", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4146, "JOAO JOSE PEREIRA", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4148, "ORLEI MAGALHAES MOREIRA", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "SUS"),
    (4149, "LUCIANA FRONTELI BELONI", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "SUS"),
    (4150, "ISABELA BATISTANUNES", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4151, "ADRIANA MOREIRA CORSINI", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "SUS"),
    (4152, "VILMA DOS SANTOS", "23/02/2026", "Gastroenterologia", "DRA RAFAELA MAIA", "SUS"),
    (4153, "SABRINA XAVIER DIAS", "23/02/2026", "Psiquiatria", "DRA MAÍRA G MELO", "SUS"),
    (4154, "MARTINS FERREIRA CHAVES", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4155, "ANGELA JULHO DA SILVA", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4156, "LUCICLEIDE LOPES DA SILVA", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4164, "ANDREIA ARRUDA", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4165, "VERA LUCIA WOZINSKI", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4170, "NIVERSINA GONCALVES DOS PASSOS", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4172, "EDUARDO CARVALHO DE SOUZA", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4175, "MARIA MARQUES RODRIGUES PESSOA", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
    (4176, "ENI AMARAL DA SILVA", "24/02/2026", "Gastroenterologia", "DRA ALICE FERREIRA", "SUS"),
]

def formatar_cpf(cpf_raw):
    if not cpf_raw:
        return ""
    digits = ''.join(c for c in str(cpf_raw) if c.isdigit())
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return str(cpf_raw)

def formatar_sus(sus_raw):
    if not sus_raw:
        return ""
    digits = ''.join(c for c in str(sus_raw) if c.isdigit())
    if len(digits) == 15:
        return f"{digits[:3]} {digits[3:7]} {digits[7:11]} {digits[11:]}"
    return str(sus_raw)

print(f"Buscando {len(atendimentos)} atendimentos na API...")
rows = []
for (id_atend, nome, data, proc, prof, operadora) in atendimentos:
    try:
        atend = api.atendimentos.detalhar(id_atend)
        id_pac = atend.get("id_paciente")
        cpf, sus = "", ""
        if id_pac:
            pac = api.pacientes.detalhar(id_pac)
            cadastro = pac.get("cadastro") or {}
            pu = pac.get("paciente_unidade") or {}
            cpf = formatar_cpf(cadastro.get("st_cpf"))
            sus = formatar_sus(pu.get("st_sus"))
        rows.append((id_atend, nome, data, proc, operadora, prof, cpf, sus))
        print(f"  [{id_atend}] {nome}: CPF={cpf or '-'} | SUS={sus or '-'}")
    except Exception as e:
        rows.append((id_atend, nome, data, proc, operadora, prof, "", ""))
        print(f"  [ERRO {id_atend}] {nome}: {e}")

# ── ESTILOS ──────────────────────────────────────────────────────────────────
AZUL_HEADER   = "1B3A6B"
AZUL_CLARO    = "D6E4F0"
VERDE         = "1E8449"
AMARELO_WARN  = "FFF3CD"
VERMELHO_WARN = "FADBD8"
CINZA_LINHA   = "F2F2F2"
BRANCO        = "FFFFFF"

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── WORKBOOK ─────────────────────────────────────────────────────────────────
wb = Workbook()

# ── ABA 1: TODOS OS ATENDIMENTOS ─────────────────────────────────────────────
ws = wb.active
ws.title = "Todos os Atendimentos"

# Título
ws.merge_cells("A1:H1")
ws["A1"] = "IRB PRIME CARE — CPF e Cartão SUS dos Atendimentos — Fevereiro 2026"
ws["A1"].font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
ws["A1"].fill = fill(AZUL_HEADER)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:H2")
ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(rows)} pacientes"
ws["A2"].font = Font(italic=True, color="555555", size=9, name="Calibri")
ws["A2"].fill = fill("EAF0FB")
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 16

# Cabeçalho
headers = ["ID Atend.", "Paciente", "Data", "Especialidade", "Operadora", "Profissional", "CPF", "Cartão SUS"]
col_widths = [10, 42, 12, 18, 12, 24, 18, 22]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cell.fill = fill(AZUL_HEADER)
    cell.alignment = center()
    cell.border = border_thin()
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[3].height = 20

# Dados
for r_idx, (id_atend, nome, data, proc, operadora, prof, cpf, sus) in enumerate(rows, 4):
    tem_cpf = bool(cpf)
    tem_sus = bool(sus)
    sem_ambos = not tem_cpf and not tem_sus

    bg = VERMELHO_WARN if sem_ambos else (CINZA_LINHA if r_idx % 2 == 0 else BRANCO)

    valores = [id_atend, nome, data, proc, operadora, prof,
               cpf if cpf else "NÃO CADASTRADO",
               sus if sus else "NÃO CADASTRADO"]

    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=r_idx, column=col, value=val)
        cell.fill = fill(bg)
        cell.border = border_thin()
        cell.font = cell_font(size=10)

        if col == 1:  # ID
            cell.alignment = center()
        elif col in (3, 5):  # Data, Operadora
            cell.alignment = center()
        else:
            cell.alignment = left()

        # CPF e SUS sem dado ficam em vermelho
        if col == 7 and not tem_cpf:
            cell.font = Font(color="C0392B", size=10, italic=True, name="Calibri")
        if col == 8 and not tem_sus:
            cell.font = Font(color="C0392B", size=10, italic=True, name="Calibri")

    ws.row_dimensions[r_idx].height = 16

# Freeze e filtro
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:H{3 + len(rows)}"

# ── ABA 2: SEM DADOS ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Sem CPF e SUS")

ws2.merge_cells("A1:H1")
ws2["A1"] = "PACIENTES SEM CPF E SEM CARTÃO SUS CADASTRADOS"
ws2["A1"].font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
ws2["A1"].fill = fill("C0392B")
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 26

ws2.merge_cells("A2:H2")
ws2["A2"] = "Estes pacientes precisam ser contactados para atualização cadastral"
ws2["A2"].font = Font(italic=True, color="7B241C", size=9, name="Calibri")
ws2["A2"].fill = fill("FADBD8")
ws2["A2"].alignment = center()
ws2.row_dimensions[2].height = 15

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cell.fill = fill("C0392B")
    cell.alignment = center()
    cell.border = border_thin()
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.row_dimensions[3].height = 20

sem_dados = [(a, b, c, d, e, f, g, h) for (a, b, c, d, e, f, g, h) in rows if not g and not h]
for r_idx, (id_atend, nome, data, proc, operadora, prof, cpf, sus) in enumerate(sem_dados, 4):
    bg = VERMELHO_WARN if r_idx % 2 == 0 else "FEF9E7"
    valores = [id_atend, nome, data, proc, operadora, prof, "NÃO CADASTRADO", "NÃO CADASTRADO"]
    for col, val in enumerate(valores, 1):
        cell = ws2.cell(row=r_idx, column=col, value=val)
        cell.fill = fill(bg)
        cell.border = border_thin()
        cell.alignment = center() if col in (1, 3, 5) else left()
        if col in (7, 8):
            cell.font = Font(color="C0392B", bold=True, size=10, name="Calibri")
        else:
            cell.font = cell_font(size=10)
    ws2.row_dimensions[r_idx].height = 16

ws2.freeze_panes = "A4"

# ── ABA 3: POR PROFISSIONAL ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Por Profissional")

# Agrupar por profissional
from collections import defaultdict
por_prof = defaultdict(list)
for row in rows:
    por_prof[row[5]].append(row)

prof_cores = {
    "DRA RAFAELA MAIA":    "1B6CA8",
    "DR ANGELO CAMPOS":    "1E8449",
    "DRA BEATRIZ ANTUNES": "8E44AD",
    "DRA ANA TRAJANO":     "D35400",
    "DRA PATRICIA BENTO":  "C0392B",
    "DRA MAÍRA G MELO":    "17A589",
    "DRA ALICE FERREIRA":  "2E4053",
}

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 42
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 12
ws3.column_dimensions["F"].width = 18
ws3.column_dimensions["G"].width = 22

current_row = 1
for prof_nome, prof_rows in sorted(por_prof.items()):
    cor = prof_cores.get(prof_nome, "2C3E50")
    total = len(prof_rows)
    com_cpf = sum(1 for r in prof_rows if r[6])
    com_sus = sum(1 for r in prof_rows if r[7])
    sem = sum(1 for r in prof_rows if not r[6] and not r[7])

    # Título do grupo
    ws3.merge_cells(f"A{current_row}:G{current_row}")
    cell = ws3.cell(row=current_row, column=1,
                    value=f"  {prof_nome}  |  {total} pacientes  |  CPF: {com_cpf}  |  SUS: {com_sus}  |  Sem dados: {sem}")
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.fill = fill(cor)
    cell.alignment = left()
    ws3.row_dimensions[current_row].height = 22
    current_row += 1

    # Sub-cabeçalho
    sub_headers = ["ID", "Paciente", "Data", "Especialidade", "Operadora", "CPF", "Cartão SUS"]
    sub_widths_cols = [1, 2, 3, 4, 5, 6, 7]
    for col, h in enumerate(sub_headers, 1):
        cell = ws3.cell(row=current_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill = fill(cor)
        cell.alignment = center()
        cell.border = border_thin()
    ws3.row_dimensions[current_row].height = 16
    current_row += 1

    for i, (id_atend, nome, data, proc, operadora, _, cpf, sus) in enumerate(prof_rows):
        bg = CINZA_LINHA if i % 2 == 0 else BRANCO
        if not cpf and not sus:
            bg = VERMELHO_WARN
        valores = [id_atend, nome, data, proc, operadora,
                   cpf or "NÃO CADASTRADO", sus or "NÃO CADASTRADO"]
        for col, val in enumerate(valores, 1):
            cell = ws3.cell(row=current_row, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border_thin()
            cell.alignment = center() if col in (1, 3, 5) else left()
            if col == 6 and not cpf:
                cell.font = Font(color="C0392B", size=9, italic=True, name="Calibri")
            elif col == 7 and not sus:
                cell.font = Font(color="C0392B", size=9, italic=True, name="Calibri")
            else:
                cell.font = Font(size=9, name="Calibri")
        ws3.row_dimensions[current_row].height = 15
        current_row += 1

    current_row += 1  # espaço entre grupos

ws3.freeze_panes = "A1"

# ── SALVAR ────────────────────────────────────────────────────────────────────
output_path = "/Users/saraiva/Documents/IRB/CPF_SUS_Fevereiro2026.xlsx"
wb.save(output_path)
print(f"\nPlanilha salva em: {output_path}")
print(f"Total de linhas: {len(rows)} | Sem dados: {len(sem_dados)}")
