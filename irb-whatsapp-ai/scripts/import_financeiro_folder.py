#!/usr/bin/env python3
from __future__ import annotations

import csv
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FINANCEIRO_DIR = ROOT / "Financeiro"
IMPORT_TAG = "[IMPORT_FINANCEIRO]"
SEED_TAG = "[SEED_FINANCEIRO]"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def slug_code(value: str, prefix: str, existing: set[str], limit: int = 20) -> str:
    base = normalize_text(value)
    base = re.sub(r"[^A-Z0-9]+", "", base)[: max(1, limit - len(prefix) - 1)] or "GEN"
    code = f"{prefix}-{base}"[:limit]
    if code not in existing:
      existing.add(code)
      return code
    index = 2
    while True:
      suffix = str(index)
      candidate = f"{prefix}-{base[: max(1, limit - len(prefix) - len(suffix) - 1)]}{suffix}"[:limit]
      if candidate not in existing:
        existing.add(candidate)
        return candidate
      index += 1


def parse_amount(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value) * 100))
    text = str(value).strip()
    if not text or text in {"-", "N/A"}:
        return None
    text = re.sub(r"[^0-9,.\-]", "", text)
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]
    if text in {"", "-", "."}:
        return None
    try:
        return int(round(float(text) * 100))
    except ValueError:
        digits = re.sub(r"[^0-9]", "", text)
        if not digits:
            return None
        if len(digits) == 1:
            normalized = f"0.0{digits}"
        elif len(digits) == 2:
            normalized = f"0.{digits}"
        else:
            normalized = f"{digits[:-2]}.{digits[-2:]}"
        return int(round(float(normalized) * 100))


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    match = re.search(r"(\d+)", str(value))
    return int(match.group(1)) if match else None


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text or normalize_text(text) in {"PREVISAO", "REFINANCIADO", "N/A", "12X", "6X"}:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def sql_today() -> date:
    return datetime.now().date()


def map_status(raw_status: Any, due_date: date | None, payment_date: date | None) -> str:
    status = normalize_text(raw_status)
    if "PAGO" in status or payment_date:
        return "paid"
    if "APROV" in status:
        return "approved"
    if "CANCEL" in status or "RETIRADO" in status or "REFINANCIADO" in status:
        return "cancelled"
    if due_date and due_date < sql_today():
        return "overdue"
    return "pending"


def infer_payment_method(*values: Any) -> str | None:
    text = normalize_text(" ".join(str(v or "") for v in values))
    if "PIX" in text:
        return "pix"
    if "BOLETO" in text:
        return "boleto"
    if "DEBITO AUTOMATICO" in text:
        return "debito_automatico"
    if "TED" in text:
        return "ted"
    if "TRANSFER" in text:
        return "transferencia"
    if "CARTAO" in text:
        return "cartao"
    return None


def find_header_row(ws, marker: str) -> int | None:
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        values = [normalize_text(v) for v in row if v not in (None, "")]
        if marker in values:
            return idx
    return None


@dataclass
class LookupMaps:
    cost_centers: dict[str, str]
    chart_accounts: dict[str, str]
    suppliers_by_doc: dict[str, str]
    suppliers_by_name: dict[str, str]
    bank_accounts: dict[str, str]
    insurance_providers: dict[str, str]
    users: list[str]
    cost_codes: set[str]
    chart_codes: set[str]


class Importer:
    def __init__(self, conn: psycopg.Connection):
        self.conn = conn
        self.cur = conn.cursor()
        self.lookups = self.load_lookups()

    def load_lookups(self) -> LookupMaps:
        cost_centers = {}
        cost_codes = set()
        self.cur.execute("select id, code, name from cost_centers")
        for row_id, code, name in self.cur.fetchall():
            cost_centers[normalize_text(name)] = str(row_id)
            cost_centers[normalize_text(code)] = str(row_id)
            cost_codes.add(code)

        chart_accounts = {}
        chart_codes = set()
        self.cur.execute("select id, code, name from chart_of_accounts")
        for row_id, code, name in self.cur.fetchall():
            chart_accounts[normalize_text(name)] = str(row_id)
            chart_accounts[normalize_text(code)] = str(row_id)
            chart_codes.add(code)

        suppliers_by_doc = {}
        suppliers_by_name = {}
        self.cur.execute("select id, coalesce(cnpj, cpf), legal_name from suppliers")
        for row_id, doc, name in self.cur.fetchall():
            if doc:
                suppliers_by_doc[normalize_text(doc)] = str(row_id)
            suppliers_by_name[normalize_text(name)] = str(row_id)

        bank_accounts = {}
        self.cur.execute("select id, bank_name, nickname from bank_accounts")
        for row_id, bank_name, nickname in self.cur.fetchall():
            if bank_name:
                bank_accounts[normalize_text(bank_name)] = str(row_id)
            if nickname:
                bank_accounts[normalize_text(nickname)] = str(row_id)

        insurance_providers = {}
        self.cur.execute("select id, code, name from insurance_providers")
        for row_id, code, name in self.cur.fetchall():
            insurance_providers[normalize_text(name)] = str(row_id)
            if code:
                insurance_providers[normalize_text(code)] = str(row_id)

        self.cur.execute("select id from users order by created_at limit 5")
        users = [str(row[0]) for row in self.cur.fetchall()]

        return LookupMaps(
            cost_centers=cost_centers,
            chart_accounts=chart_accounts,
            suppliers_by_doc=suppliers_by_doc,
            suppliers_by_name=suppliers_by_name,
            bank_accounts=bank_accounts,
            insurance_providers=insurance_providers,
            users=users,
            cost_codes=cost_codes,
            chart_codes=chart_codes,
        )

    @property
    def default_user_id(self) -> str | None:
        return self.lookups.users[0] if self.lookups.users else None

    def cleanup_previous_imports(self) -> None:
        statements = [
            """
            delete from bank_transactions
            where account_payable_id in (
              select id from accounts_payable where notes like %s or notes like %s
            )
            or account_receivable_id in (
              select id from accounts_receivable where notes like %s or notes like %s
            )
            """,
            "delete from payment_approvals where notes like %s or notes like %s",
            "delete from bank_transactions where external_ref like %s or external_ref like %s",
            "delete from receivable_payments where notes like %s or notes like %s",
            "delete from receivable_installments where account_receivable_id in (select id from accounts_receivable where notes like %s or notes like %s)",
            "delete from accounts_receivable where notes like %s or notes like %s",
            "delete from reimbursement_items where reimbursement_request_id in (select id from reimbursement_requests where notes like %s or notes like %s)",
            "delete from reimbursement_requests where notes like %s or notes like %s",
            "delete from transport_vouchers where notes like %s or notes like %s",
            "delete from cash_flow_snapshots where notes like %s or notes like %s",
            "delete from credit_card_purchases where description like %s or description like %s",
            "delete from accounts_payable where notes like %s or notes like %s",
        ]
        for statement in statements:
            placeholder_count = statement.count("%s")
            params = []
            for _ in range(placeholder_count // 2):
                params.extend((f"{IMPORT_TAG}%", f"{SEED_TAG}%"))
            self.cur.execute(statement, tuple(params))

    def ensure_cost_center(self, raw_name: Any) -> str | None:
        if raw_name is None or str(raw_name).strip() == "":
            return None
        aliases = {
            "CLINICA": "IRB PRIME CARE",
            "CLINICA/PROJETOS": "PROJETOS",
            "PROJETOS": "PROJETOS",
            "RONDONIA": "RONDONIA",
            "DIRETORIA": "DIRETORIA",
            "ADM": "ADMINISTRATIVO",
            "NARDINI": "NARDINI",
            "BRAGANCA": "BRAGANCA",
            "PARAGUACU": "PARAGUACU",
            "SAMU - MG": "SAMU MINAS GERAIS",
            "SAMU MG": "SAMU MINAS GERAIS",
            "TI": "TECNOLOGIA",
            "FINANCEIRO": "FINANCEIRO",
            "FATURAMENTO": "FATURAMENTO",
        }
        normalized = normalize_text(raw_name)
        normalized = aliases.get(normalized, normalized)
        existing = self.lookups.cost_centers.get(normalized)
        if existing:
            return existing
        code = slug_code(normalized, "CC", self.lookups.cost_codes)
        self.cur.execute(
            """
            insert into cost_centers (code, name, description, is_active)
            values (%s, %s, %s, true)
            returning id
            """,
            (code, str(raw_name).strip(), f"{IMPORT_TAG} centro de custo importado"),
        )
        row_id = str(self.cur.fetchone()[0])
        self.lookups.cost_centers[normalized] = row_id
        return row_id

    def ensure_chart_account(self, raw_name: Any, account_type: str = "expense") -> str | None:
        if raw_name is None or str(raw_name).strip() == "":
            return None
        aliases = {
            "VIAGEM": "VIAGENS E DESLOCAMENTOS",
            "TAXAS / LICENCAS": "TAXAS / LICENCAS",
            "TAXAS / LICENÇAS": "TAXAS / LICENCAS",
            "PRESTADORES DE SERVICOS": "SERVICOS PROFISSIONAIS",
            "PRESTADORES DE SERVIÇOS": "SERVICOS PROFISSIONAIS",
            "TELEFONE E INTERNET": "UTILIDADES",
            "CARTAO DE CREDITO": "FINANCEIRO",
            "JUROS": "JUROS E MULTAS",
            "TARIFA": "TARIFAS BANCARIAS",
            "INSUMOS": "MATERIAIS E INSUMOS",
            "CONTABILIDADE": "SERVICOS PROFISSIONAIS",
            "TI - DESPESAS/SERVICOS": "SERVICOS PROFISSIONAIS",
        }
        normalized = normalize_text(raw_name)
        normalized = aliases.get(normalized, normalized)
        existing = self.lookups.chart_accounts.get(normalized)
        if existing:
            return existing
        code = slug_code(normalized, "CH", self.lookups.chart_codes)
        self.cur.execute(
            """
            insert into chart_of_accounts (code, name, type, is_active)
            values (%s, %s, %s, true)
            returning id
            """,
            (code, str(raw_name).strip(), account_type),
        )
        row_id = str(self.cur.fetchone()[0])
        self.lookups.chart_accounts[normalized] = row_id
        return row_id

    def ensure_supplier(self, name: Any, document: Any, notes: str) -> str | None:
        if name is None or str(name).strip() == "":
            return None
        normalized_name = normalize_text(name)
        normalized_doc = normalize_text(document) if document else ""
        doc_digits = re.sub(r"\D", "", str(document or ""))
        if normalized_doc and normalized_doc in self.lookups.suppliers_by_doc:
            return self.lookups.suppliers_by_doc[normalized_doc]
        if doc_digits:
            self.cur.execute(
                """
                select id, coalesce(cnpj, cpf), legal_name
                from suppliers
                where regexp_replace(coalesce(cnpj, cpf, ''), '\\D', '', 'g') = %s
                limit 1
                """,
                (doc_digits,),
            )
            row = self.cur.fetchone()
            if row:
                row_id = str(row[0])
                if normalized_doc:
                    self.lookups.suppliers_by_doc[normalized_doc] = row_id
                self.lookups.suppliers_by_name[normalized_name] = row_id
                return row_id
        if normalized_name in self.lookups.suppliers_by_name:
            return self.lookups.suppliers_by_name[normalized_name]

        cnpj = None
        cpf = None
        doc_text = str(document or "").strip()
        if re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", doc_text):
            cnpj = re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", doc_text).group(0)
        elif re.search(r"\d{3}\.\d{3}\.\d{3}-\d{2}", doc_text):
            cpf = re.search(r"\d{3}\.\d{3}\.\d{3}-\d{2}", doc_text).group(0)
        elif re.fullmatch(r"\d{11}", re.sub(r"\D", "", doc_text or "")):
            digits = re.sub(r"\D", "", doc_text)
            cpf = f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"

        self.cur.execute(
            """
            insert into suppliers (cnpj, cpf, legal_name, trade_name, notes, is_active)
            values (%s, %s, %s, %s, %s, true)
            returning id
            """,
            (cnpj, cpf, str(name).strip(), str(name).strip(), notes),
        )
        row_id = str(self.cur.fetchone()[0])
        if normalized_doc:
            self.lookups.suppliers_by_doc[normalized_doc] = row_id
        self.lookups.suppliers_by_name[normalized_name] = row_id
        return row_id

    def bank_account_for(self, raw_bank: Any) -> str | None:
        if raw_bank is None:
            return None
        text = normalize_text(raw_bank)
        for key, row_id in self.lookups.bank_accounts.items():
            if key and key in text:
                return row_id
        return None

    def ensure_insurance_provider(self, name: Any) -> str | None:
        if name is None or str(name).strip() == "":
            return None
        normalized = normalize_text(name)
        existing = self.lookups.insurance_providers.get(normalized)
        if existing:
            return existing
        code = slug_code(normalized, "CV", set(self.lookups.insurance_providers.keys()), 20)
        self.cur.execute(
            """
            insert into insurance_providers (code, name, payment_term_days, is_active)
            values (%s, %s, 30, true)
            returning id
            """,
            (code, str(name).strip()),
        )
        row_id = str(self.cur.fetchone()[0])
        self.lookups.insurance_providers[normalized] = row_id
        self.lookups.insurance_providers[normalize_text(code)] = row_id
        return row_id

    def import_convenios(self) -> int:
        path = FINANCEIRO_DIR / "Contas a Receber IRB.xlsx"
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["Cadastro Convenios"]
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if not name:
                continue
            before = self.lookups.insurance_providers.get(normalize_text(name))
            self.ensure_insurance_provider(name)
            if before is None:
                imported += 1
        return imported

    def import_accounts_receivable(self) -> int:
        path = FINANCEIRO_DIR / "Contas a Receber IRB.xlsx"
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["Contas a Receber"]
        header_row = find_header_row(ws, "PACIENTE")
        if not header_row:
            return 0
        imported = 0
        for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            patient = row[1]
            procedure = row[7]
            total_amount = parse_amount(row[10])
            if not patient or not procedure or not total_amount:
                continue
            received_amount = parse_amount(row[11]) or 0
            balance = parse_amount(row[12]) or max(total_amount - received_amount, 0)
            payment_type = "insurance" if "CONVENIO" in normalize_text(row[4]) else "particular"
            insurance_id = self.ensure_insurance_provider(row[5]) if payment_type == "insurance" else None
            cost_center_id = self.ensure_cost_center("FATURAMENTO" if payment_type == "insurance" else "CLINICA")
            status = "received" if balance <= 0 else ("overdue" if (parse_date(row[14]) and parse_date(row[14]) < sql_today()) else "pending")
            self.cur.execute(
                """
                insert into accounts_receivable (
                  patient_id, insurance_provider_id, cost_center_id, service_type, procedure_description,
                  guide_number, total_amount, received_amount, service_date, due_date, received_date,
                  status, payment_type, notes, created_by
                ) values (
                  null, %s, %s, %s, %s,
                  %s, %s, %s, %s, %s, %s,
                  %s, %s, %s, %s
                )
                """,
                (
                    insurance_id,
                    cost_center_id,
                    "medical" if "MEDICO" in normalize_text(row[3]) else "dental",
                    str(procedure).strip(),
                    str(row[6]).strip() if row[6] else None,
                    total_amount,
                    received_amount,
                    parse_date(row[9]),
                    parse_date(row[14]),
                    parse_date(row[13]) if received_amount else None,
                    status,
                    payment_type,
                    f"{IMPORT_TAG} arquivo=Contas a Receber IRB.xlsx linha={idx} paciente={patient}",
                    self.default_user_id,
                ),
            )
            imported += 1
        return imported

    def import_payables_sheet(self) -> int:
        path = FINANCEIRO_DIR / "FLUXO DE PAGAMENTO- 2026.xlsx"
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["FLUXO DE PAGAMENTOS"]
        imported = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            supplier_name = row[3]
            amount = parse_amount(row[10])
            if not supplier_name or amount is None or amount <= 0:
                continue
            description = str(row[13] or supplier_name).strip()
            issue_date = parse_date(row[6])
            due_date = parse_date(row[7]) or issue_date or sql_today()
            payment_date = parse_date(row[8])
            status = map_status(row[12], due_date, payment_date)
            if status == "cancelled":
                continue
            cost_center_id = self.ensure_cost_center(row[1])
            chart_id = self.ensure_chart_account(row[2], "expense")
            supplier_id = self.ensure_supplier(supplier_name, row[4], f"{IMPORT_TAG} fornecedor planilha fluxo")
            bank_id = self.bank_account_for(row[11])
            inss = parse_amount(row[15]) or 0
            irpj = parse_amount(row[16]) or 0
            csll = parse_amount(row[17]) or 0
            cofins = parse_amount(row[18]) or 0
            pis = parse_amount(row[19]) or 0
            iss = parse_amount(row[20]) or 0
            tarifa = parse_amount(row[21]) or 0
            juros = parse_amount(row[22]) or 0
            net_amount = amount + tarifa + juros
            gross_amount = net_amount + inss + irpj + csll + cofins + pis + iss
            self.cur.execute(
                """
                insert into accounts_payable (
                  document_number, document_type, supplier_id, cost_center_id, chart_account_id, bank_account_id,
                  description, gross_amount, net_amount, inss_retention, irpj_retention, csll_retention,
                  cofins_retention, pis_retention, iss_retention, issue_date, due_date, payment_date,
                  competence_date, status, payment_method, notes, created_by, approved_by, approved_at, paid_by
                ) values (
                  %s, %s, %s, %s, %s, %s,
                  %s, %s, %s, %s, %s, %s,
                  %s, %s, %s, %s, %s, %s,
                  %s, %s, %s, %s, %s, %s, %s, %s
                )
                returning id
                """,
                (
                    f"IMP-{row[0]}" if row[0] is not None else f"FLUXO-{idx}",
                    str(row[5]).strip() if row[5] else "N/A",
                    supplier_id,
                    cost_center_id,
                    chart_id,
                    bank_id,
                    description,
                    gross_amount,
                    net_amount,
                    inss,
                    irpj,
                    csll,
                    cofins,
                    pis,
                    iss,
                    issue_date,
                    due_date,
                    payment_date,
                    issue_date,
                    status,
                    infer_payment_method(row[11], description, row[4], row[14]),
                    f"{IMPORT_TAG} arquivo=FLUXO DE PAGAMENTO- 2026.xlsx aba=FLUXO DE PAGAMENTOS linha={idx} mes={row[9]} tipo={row[14]}",
                    self.default_user_id,
                    self.default_user_id if status in {"approved", "paid"} else None,
                    datetime.now() if status in {"approved", "paid"} else None,
                    self.default_user_id if status == "paid" else None,
                ),
            )
            imported += 1
        return imported

    def import_daily_approvals(self) -> int:
        path = FINANCEIRO_DIR / "FLUXO DE PAGAMENTO- 2026.xlsx"
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["PAGAMENTO DIARIO (2)"]
        header_row = find_header_row(ws, "FORNECEDOR/PRESTADOR")
        if not header_row:
            return 0
        imported = 0
        for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            supplier_name = row[1]
            amount = parse_amount(row[7])
            if not supplier_name or amount is None or amount <= 0:
                continue
            issue_date = parse_date(row[4])
            due_date = parse_date(row[5]) or issue_date or sql_today()
            supplier_id = self.ensure_supplier(supplier_name, None, f"{IMPORT_TAG} fornecedor aprovacao diaria")
            cost_center_id = self.ensure_cost_center(row[2])
            chart_id = self.ensure_chart_account("Aprovação diária", "expense")
            status = map_status(row[8], due_date, None)
            self.cur.execute(
                """
                insert into accounts_payable (
                  document_number, document_type, supplier_id, cost_center_id, chart_account_id,
                  description, gross_amount, net_amount, issue_date, due_date, status,
                  payment_method, notes, created_by
                ) values (
                  %s, %s, %s, %s, %s,
                  %s, %s, %s, %s, %s, %s,
                  %s, %s, %s
                )
                returning id
                """,
                (
                    f"APD-{idx}",
                    str(row[3]).strip() if row[3] else "N/A",
                    supplier_id,
                    cost_center_id,
                    chart_id,
                    str(row[9] or supplier_name).strip(),
                    amount,
                    amount,
                    issue_date,
                    due_date,
                    status,
                    infer_payment_method(row[9], row[10]),
                    f"{IMPORT_TAG} arquivo=FLUXO DE PAGAMENTO- 2026.xlsx aba=PAGAMENTO DIARIO (2) linha={idx}",
                    self.default_user_id,
                ),
            )
            payable_id = str(self.cur.fetchone()[0])
            self.cur.execute(
                """
                insert into payment_approvals (
                  account_payable_id, requested_by, requested_at, status, notes
                ) values (%s, %s, now(), %s, %s)
                """,
                (
                    payable_id,
                    self.default_user_id,
                    "approved" if status == "approved" else "pending",
                    f"{IMPORT_TAG} aprovação diária linha={idx}",
                ),
            )
            imported += 1
        return imported

    def import_credit_card(self) -> int:
        path = FINANCEIRO_DIR / "FLUXO DE PAGAMENTO- 2026.xlsx"
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["COMPRAS CARTÃO DE CRÉDITO"]
        imported = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            merchant = row[3]
            amount = parse_amount(row[10])
            if not merchant or amount is None or amount <= 0:
                continue
            installments = parse_int(row[8]) or 1
            self.cur.execute(
                """
                insert into credit_card_purchases (
                  card_last_digits, card_holder, merchant_name, purchase_date, total_amount,
                  installments, installment_amount, current_installment, cost_center_id,
                  chart_account_id, description, status
                ) values (
                  %s, %s, %s, %s, %s,
                  %s, %s, %s, %s,
                  %s, %s, %s
                )
                """,
                (
                    "3646",
                    "IRB",
                    str(merchant).strip(),
                    parse_date(row[6]) or sql_today(),
                    amount,
                    installments,
                    amount,
                    1,
                    self.ensure_cost_center(row[1] or "CLINICA"),
                    self.ensure_chart_account(row[2] or "Cartão de crédito", "expense"),
                    f"{str(row[12] or '').strip()} {IMPORT_TAG} arquivo=FLUXO DE PAGAMENTO- 2026.xlsx aba=COMPRAS CARTÃO DE CRÉDITO linha={idx}".strip(),
                    "paid" if "PAGO" in normalize_text(row[11]) else "active",
                ),
            )
            imported += 1
        return imported

    def import_reimbursement(self) -> int:
        path = FINANCEIRO_DIR / "FORMULÁRIO DE REEMBOLSO - Viagem 02032026 a 05032026.xlsx"
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        data = {}
        items = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            first = row[0]
            if isinstance(first, str) and ":" in first:
                key, value = first.split(":", 1)
                data[normalize_text(key)] = value.strip()
            if idx == 12 and row[7]:
                text = str(row[7])
                if ":" in text:
                    _, value = text.split(":", 1)
                    data["DATA DE TERMINO"] = value.strip()
            if idx in (24, 25):
                expense_date = parse_date(row[0])
                expense_type = str(row[1]).strip() if row[1] else "OUTRO"
                description = str(row[3]).strip() if row[3] else None
                amount = parse_amount(row[10])
                receipt = str(row[2]).strip() if row[2] else None
                if expense_date and amount:
                    items.append((expense_date, expense_type, description, receipt, amount))

        total_amount = parse_amount(329.48) if items else 0
        employee_doc = data.get("CNPJ/CNPJ")
        employee_cpf = employee_doc if employee_doc and len(str(employee_doc).strip()) <= 14 else None
        self.cur.execute(
            """
            insert into reimbursement_requests (
              request_number, employee_name, employee_department, employee_cpf,
              trip_origin, trip_destination, trip_start_date, trip_end_date, trip_purpose,
              bank_name, bank_agency, bank_account, bank_account_type, pix_key,
              total_amount, approved_amount, status, requested_by, notes
            ) values (
              %s, %s, %s, %s,
              %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s
            )
            returning id
            """,
            (
                "REEMB-2026-03-05-FRANDIS",
                data.get("NOME DO PROFISSIONAL") or "Frandis Rafael Rodrigues Vasconcelos",
                (data.get("DEPARTAMENTO / CENTRO DE CUSTO") or "TI").strip(),
                employee_cpf,
                "Araraquara",
                "Sao Paulo",
                parse_date(data.get("DATA DE INICIO")) or date(2026, 3, 2),
                parse_date(data.get("DATA DE TERMINO")) or date(2026, 3, 5),
                data.get("FINALIDADE E/OU ITINERARIO") or "Viagem corporativa",
                data.get("BANCO") or "Bradesco",
                data.get("AGENCIA") or "2700",
                data.get("CONTA CORRENTE") or "13532",
                "corrente",
                data.get("CHAVE PIX") or data.get("CNPJ/CNPJ"),
                total_amount,
                total_amount,
                "approved",
                self.default_user_id,
                f"{IMPORT_TAG} arquivo={path.name}",
            ),
        )
        request_id = str(self.cur.fetchone()[0])
        for expense_date, expense_type, description, receipt, amount in items:
            self.cur.execute(
                """
                insert into reimbursement_items (
                  reimbursement_request_id, expense_date, expense_type, description,
                  receipt_number, amount, approved, approved_amount
                ) values (%s, %s, %s, %s, %s, %s, true, %s)
                """,
                (request_id, expense_date, expense_type.lower(), description, receipt, amount, amount),
            )
        return 1

    def import_transport_vouchers(self) -> int:
        path = FINANCEIRO_DIR / "ORDEM DE PAGAMENTO VT.xlsx"
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        imported = 0
        reference_date = date(2026, 3, 5)
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            employee = row[0]
            amount = parse_amount(row[7])
            if idx < 15 or not employee or amount is None or amount <= 0:
                continue
            self.cur.execute(
                """
                insert into transport_vouchers (
                  employee_name, employee_role, contract_type, cost_center_id,
                  monthly_amount, reference_month, status, notes
                ) values (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    str(employee).strip(),
                    "Colaborador CLT",
                    "clt",
                    self.ensure_cost_center("RH"),
                    amount,
                    reference_date.strftime("%Y-%m"),
                    "pending",
                    f"{IMPORT_TAG} arquivo={path.name} linha={idx}",
                ),
            )
            imported += 1
        return imported

    def import_cash_flow_snapshots(self) -> int:
        path = FINANCEIRO_DIR / "Fluxo de Caixa (1)(Fluxo de caixa - Fevereiro-26).csv"
        with path.open(encoding="latin1", newline="") as fp:
            rows = list(csv.reader(fp, delimiter=";"))

        headers = rows[0][1:21]
        opening = rows[2][1:21]
        closing = rows[105][1:21]
        imported = 0
        for idx, day in enumerate(headers):
            if not day:
                continue
            snapshot_date = parse_date(day)
            opening_balance = parse_amount(opening[idx]) or 0
            closing_balance = parse_amount(closing[idx]) or opening_balance
            delta = closing_balance - opening_balance
            credits = delta if delta > 0 else 0
            debits = -delta if delta < 0 else 0
            self.cur.execute(
                """
                insert into cash_flow_snapshots (
                  snapshot_date, cost_center_id, opening_balance, total_credits,
                  total_debits, closing_balance, is_projected, notes, generated_by
                ) values (%s, null, %s, %s, %s, %s, false, %s, %s)
                """,
                (
                    snapshot_date,
                    opening_balance,
                    credits,
                    debits,
                    closing_balance,
                    f"{IMPORT_TAG} arquivo={path.name}",
                    self.default_user_id,
                ),
            )
            imported += 1
        return imported


def main() -> None:
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise SystemExit("DATABASE_URL nao definido")
    if not FINANCEIRO_DIR.exists():
        raise SystemExit(f"Pasta nao encontrada: {FINANCEIRO_DIR}")

    with psycopg.connect(database_url) as conn:
        importer = Importer(conn)
        importer.cleanup_previous_imports()
        results = {
            "convenios": importer.import_convenios(),
            "contas_receber": importer.import_accounts_receivable(),
            "contas_pagar": importer.import_payables_sheet(),
            "aprovacoes_diarias": importer.import_daily_approvals(),
            "cartao_credito": importer.import_credit_card(),
            "reembolsos": importer.import_reimbursement(),
            "vale_transporte": importer.import_transport_vouchers(),
            "fluxo_caixa": importer.import_cash_flow_snapshots(),
        }
        conn.commit()

    print("Importacao concluida:")
    for key, value in results.items():
        print(f"- {key}: {value}")


if __name__ == "__main__":
    main()
